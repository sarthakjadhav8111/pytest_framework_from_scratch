import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,columnno).value=data
    workbook.save(file)

def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    cell = sheet.cell(row=rownum, column=columnno)
    cell.fill = green_fill
    workbook.save(file)

def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    cell = sheet.cell(row=rownum, column=columnno)
    cell.fill = red_fill
    workbook.save(file)

